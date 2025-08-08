#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Shop Management 모듈
- ERPia 매장(거래처) 정보 관리
- 레거시 ShopManager와 동일한 기능
"""

from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, send_file
from app.common.models import db, ErpiaCustomer, Code
from sqlalchemy import or_, and_
import logging
from datetime import datetime
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

# Blueprint 생성
shop_bp = Blueprint('shop', __name__, url_prefix='/shop')

@shop_bp.route('/')
@shop_bp.route('/index')
def index():
    """매장(거래처) 관리 메인 페이지"""
    # 세션 체크
    if not session.get('member_seq'):
        return redirect(url_for('auth.login'))
    
    current_company_id = session.get('current_company_id', 1)
    
    # 매장 유형 필터
    shop_type = request.args.get('ty', 'All')
    
    logger.info(f"🏪 매장 관리 페이지 접근: company_id={current_company_id}, shop_type={shop_type}")
    
    return render_template('shop/index.html', shop_type=shop_type)

@shop_bp.route('/api/shops')
def api_get_shops():
    """매장 목록 조회 API"""
    try:
        current_company_id = session.get('current_company_id', 1)
        shop_type = request.args.get('ty', 'All')
        search = request.args.get('search', '').strip()
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        
        # 기본 쿼리
        query = ErpiaCustomer.query.filter_by(company_id=current_company_id)
        
        # 매장 유형 필터링
        if shop_type == 'Shop':
            query = query.filter_by(shop_yn='Y')
        elif shop_type == 'NoShop':
            query = query.filter_by(shop_yn='N')
        elif shop_type == 'NotUsed':
            query = query.filter_by(shop_yn='NU')  # 매장사용안함
        elif shop_type == 'Nothing':
            query = query.filter(ErpiaCustomer.shop_yn.is_(None))
        
        # 검색 필터링
        if search:
            query = query.filter(
                or_(
                    ErpiaCustomer.customer_name.like(f'%{search}%'),
                    ErpiaCustomer.customer_code.like(f'%{search}%'),
                    ErpiaCustomer.our_manager.like(f'%{search}%'),
                    ErpiaCustomer.address1.like(f'%{search}%')
                )
            )
        
        # 페이징
        total = query.count()
        shops = query.order_by(ErpiaCustomer.customer_name.asc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        return jsonify({
            'success': True,
            'data': [shop.to_dict() for shop in shops.items],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': shops.pages,
                'has_prev': shops.has_prev,
                'has_next': shops.has_next
            }
        })
    except Exception as e:
        logger.error(f"❌ 매장 목록 조회 실패: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@shop_bp.route('/api/shop/<int:shop_seq>')
def api_get_shop_info(shop_seq):
    """매장 상세 정보 조회 API"""
    try:
        current_company_id = session.get('current_company_id', 1)
        
        shop = ErpiaCustomer.query.filter_by(
            seq=shop_seq,
            company_id=current_company_id
        ).first()
        
        if not shop:
            return jsonify({'success': False, 'message': '매장을 찾을 수 없습니다.'}), 404
        
        return jsonify({
            'success': True,
            'data': shop.to_dict()
        })
    except Exception as e:
        logger.error(f"❌ 매장 정보 조회 실패: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@shop_bp.route('/api/shop/<int:shop_seq>', methods=['PUT'])
def api_update_shop(shop_seq):
    """매장 정보 수정 API"""
    try:
        current_company_id = session.get('current_company_id', 1)
        member_id = session.get('member_id', 'admin')
        
        shop = ErpiaCustomer.query.filter_by(
            seq=shop_seq,
            company_id=current_company_id
        ).first()
        
        if not shop:
            return jsonify({'success': False, 'message': '매장을 찾을 수 없습니다.'}), 404
        
        data = request.get_json()
        
        # 수정 가능한 필드들 업데이트 (동적으로 처리)
        # 시스템 필드는 제외하고 사용자가 수정할 수 있는 필드만 처리
        exclude_fields = {
            'seq', 'company_id', 'customer_code', 'ins_user', 'ins_date', 'upt_user', 'upt_date'
        }
        
        # ErpiaCustomer 모델의 모든 컬럼 가져오기
        from sqlalchemy import inspect
        mapper = inspect(ErpiaCustomer)
        updatable_fields = []
        
        for column in mapper.columns:
            field_name = column.name
            if field_name not in exclude_fields:
                updatable_fields.append(field_name)
        
        # 데이터에서 수정 가능한 필드들만 업데이트
        for field in updatable_fields:
            if field in data:
                setattr(shop, field, data[field])
        
        # 수정 정보 업데이트
        shop.upt_user = member_id
        from datetime import datetime
        shop.upt_date = datetime.utcnow()
        
        db.session.commit()
        
        logger.info(f"✅ 매장 정보 수정 완료: seq={shop_seq}, name={shop.customer_name}")
        
        return jsonify({
            'success': True,
            'message': '매장 정보가 수정되었습니다.',
            'data': shop.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ 매장 정보 수정 실패: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@shop_bp.route('/api/classifications')
def api_get_classifications():
    """매장 분류정보 조회 API - CST 코드 관리와 완전 연동"""
    try:
        classifications = {}
        
        # CST 그룹 찾기
        cst_group = Code.query.filter_by(code='CST', depth=0).first()
        if not cst_group:
            logger.warning("❌ CST 그룹을 찾을 수 없습니다.")
            return jsonify({'success': True, 'data': {}})
        
        logger.info(f"🔍 CST 그룹 찾음: {cst_group.code_name} (seq: {cst_group.seq})")
        
        # CST 하위의 모든 분류 그룹들을 동적으로 조회 (depth=1)
        classification_groups = Code.query.filter_by(
            parent_seq=cst_group.seq, 
            depth=1
        ).order_by(Code.sort.asc()).all()
        
        logger.info(f"🏷️ 발견된 분류 그룹 수: {len(classification_groups)}")
        
        for group in classification_groups:
            # 각 분류 그룹의 하위 코드들 조회 (depth=2)
            sub_codes = Code.query.filter_by(
                parent_seq=group.seq,
                depth=2
            ).order_by(Code.sort.asc()).all()
            
            # 분류 그룹 키는 소문자로 통일
            group_key = group.code.lower()
            classifications[group_key] = {
                'group_info': {
                    'code': group.code,
                    'name': group.code_name,
                    'seq': group.seq,
                    'sort': group.sort
                },
                'codes': [code.to_dict() for code in sub_codes]
            }
            
            logger.info(f"  📂 {group.code}({group.code_name}): {len(sub_codes)}개 하위 코드")
        
        # 레거시 호환을 위해 기존 키 형태도 유지
        legacy_mapping = {
            'dis': '유통', 'ch': '채널', 'sl': '매출', 'ty': '형태', 'grd': '등급',
            'bz': '브랜드존', 'nz': '뉴나브랜드조닝', 'rg': '지역', 'fg': '가결산구분값'
        }
        
        # 결과에 메타정보 추가
        result = {
            'success': True,
            'data': classifications,
            'meta': {
                'total_groups': len(classification_groups),
                'cst_group': {
                    'seq': cst_group.seq,
                    'name': cst_group.code_name
                },
                'last_updated': datetime.utcnow().isoformat()
            }
        }
        
        logger.info(f"✅ 매장 분류 정보 조회 완료: {len(classifications)}개 그룹")
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"❌ 분류 정보 조회 오류: {str(e)}")
        return jsonify({
            'success': False, 
            'message': str(e),
            'data': {}
        }), 500

@shop_bp.route('/api/classifications/groups')
def api_get_classification_groups():
    """분류 그룹 목록만 조회 (UI 구성용)"""
    try:
        # CST 그룹 찾기
        cst_group = Code.query.filter_by(code='CST', depth=0).first()
        if not cst_group:
            return jsonify({'success': True, 'data': []})
        
        # CST 하위의 모든 분류 그룹들 조회
        groups = Code.query.filter_by(
            parent_seq=cst_group.seq, 
            depth=1
        ).order_by(Code.sort.asc()).all()
        
        result = []
        for group in groups:
            # 각 그룹의 하위 코드 개수 계산
            sub_count = Code.query.filter_by(parent_seq=group.seq, depth=2).count()
            
            result.append({
                'code': group.code,
                'name': group.code_name,
                'seq': group.seq,
                'sort': group.sort,
                'sub_codes_count': sub_count,
                'key': group.code.lower()  # UI에서 사용할 키
            })
        
        return jsonify({
            'success': True,
            'data': result,
            'total': len(result)
        })
        
    except Exception as e:
        logger.error(f"❌ 분류 그룹 목록 조회 오류: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@shop_bp.route('/api/stats')
def api_get_shop_stats():
    """매장 통계 정보 조회 API"""
    try:
        current_company_id = session.get('current_company_id', 1)
        
        total = ErpiaCustomer.query.filter_by(company_id=current_company_id).count()
        shops = ErpiaCustomer.query.filter_by(company_id=current_company_id, shop_yn='Y').count()
        no_shops = ErpiaCustomer.query.filter_by(company_id=current_company_id, shop_yn='N').count()
        not_used = ErpiaCustomer.query.filter_by(company_id=current_company_id, shop_yn='NU').count() # 매장사용안함
        nothing = ErpiaCustomer.query.filter(
            and_(
                ErpiaCustomer.company_id == current_company_id,
                ErpiaCustomer.shop_yn.is_(None)
            )
        ).count()
        
        return jsonify({
            'success': True,
            'data': {
                'total': total,
                'shops': shops,
                'no_shops': no_shops,
                'not_used': not_used,
                'nothing': nothing
            }
        })
    except Exception as e:
        logger.error(f"❌ 매장 통계 조회 실패: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500 

@shop_bp.route('/api/sync-erpia', methods=['POST'])
def api_sync_erpia_customers():
    """ERPia에서 매장 정보 동기화 API"""
    try:
        logger.info("🔄 ERPia 동기화 API 호출됨")
        
        # 세션 체크
        if not session.get('member_seq'):
            logger.warning("❌ 인증 실패: 세션 없음")
            return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401
        
        current_company_id = session.get('current_company_id', 1)
        member_id = session.get('member_id', 'admin')
        
        logger.info(f"✅ 세션 확인됨: company_id={current_company_id}, member_id={member_id}")
        
        # 임포트 테스트
        try:
            from app.services.erpia_client import ErpiaApiClient
            logger.info("✅ ErpiaApiClient 임포트 성공")
        except Exception as e:
            logger.error(f"❌ ErpiaApiClient 임포트 실패: {str(e)}")
            return jsonify({'success': False, 'message': f'모듈 임포트 오류: {str(e)}'}), 500
        
        try:
            from app.common.models import CompanyErpiaConfig
            logger.info("✅ CompanyErpiaConfig 임포트 성공")
        except Exception as e:
            logger.error(f"❌ CompanyErpiaConfig 임포트 실패: {str(e)}")
            return jsonify({'success': False, 'message': f'모델 임포트 오류: {str(e)}'}), 500
        
        # 회사별 ERPia 설정 조회
        try:
            erpia_config = CompanyErpiaConfig.query.filter_by(company_id=current_company_id).first()
            logger.info(f"✅ ERPia 설정 조회 완료: {erpia_config is not None}")
        except Exception as e:
            logger.error(f"❌ ERPia 설정 조회 실패: {str(e)}")
            return jsonify({'success': False, 'message': f'설정 조회 오류: {str(e)}'}), 500
        
        if not erpia_config:
            logger.warning("❌ ERPia 설정 없음")
            return jsonify({
                'success': False, 
                'message': 'ERPia 연동 설정이 없습니다. 배치 설정에서 ERPia 설정을 먼저 구성해주세요.'
            }), 400
        
        # ERPia 클라이언트 생성 테스트
        try:
            logger.info(f"🔧 ERPia 클라이언트 생성 중: company_id={current_company_id}")
            erpia_client = ErpiaApiClient(company_id=current_company_id)
            logger.info("✅ ERPia 클라이언트 생성 성공")
        except Exception as e:
            logger.error(f"❌ ERPia 클라이언트 생성 실패: {str(e)}")
            return jsonify({'success': False, 'message': f'ERPia 클라이언트 생성 오류: {str(e)}'}), 500
        
        # 간단한 연결 테스트
        try:
            logger.info("🔌 ERPia 연결 테스트 중...")
            test_result = erpia_client.test_connection()
            logger.info(f"✅ ERPia 연결 테스트 결과: {test_result}")
        except Exception as e:
            logger.error(f"❌ ERPia 연결 테스트 실패: {str(e)}")
            return jsonify({'success': False, 'message': f'ERPia 연결 실패: {str(e)}'}), 500
        
        # 실제 데이터 동기화 시작
        try:
            logger.info("📡 ERPia에서 고객 데이터 가져오는 중...")
            
            # 분류 코드 매핑 준비 (CST 하위 분류들)
            classification_mapping = {}
            try:
                cst_group = Code.query.filter_by(code='CST', depth=0).first()
                if cst_group:
                    classification_groups = Code.query.filter_by(
                        parent_seq=cst_group.seq, 
                        depth=1
                    ).all()
                    
                    for group in classification_groups:
                        group_codes = Code.query.filter_by(
                            parent_seq=group.seq,
                            depth=2
                        ).all()
                        
                        # 코드별 매핑 딕셔너리 생성
                        group_key = group.code.lower()
                        classification_mapping[group_key] = {}
                        for code in group_codes:
                            classification_mapping[group_key][code.code] = code.code_name
                    
                    logger.info(f"📋 분류 매핑 준비 완료: {list(classification_mapping.keys())}")
                else:
                    logger.warning("⚠️ CST 분류 그룹을 찾을 수 없습니다.")
            except Exception as e:
                logger.error(f"❌ 분류 매핑 준비 실패: {str(e)}")
                classification_mapping = {}
            
            # 날짜 범위 설정 (최근 1년간)
            from datetime import datetime, timedelta
            end_date = datetime.now()
            start_date = end_date - timedelta(days=365)  # 1년 전
            
            start_date_str = start_date.strftime('%Y%m%d')
            end_date_str = end_date.strftime('%Y%m%d')
            
            logger.info(f"📅 조회 기간: {start_date_str} ~ {end_date_str}")
            logger.info(f"🏢 대상 회사: {current_company_id} ({'에이원' if current_company_id == 1 else '에이원월드'})")
            
            # ERPia에서 고객 정보 가져오기
            customers_data = erpia_client.fetch_customers(
                start_date=start_date_str,
                end_date=end_date_str
            )
            
            if not customers_data:
                logger.warning("⚠️ ERPia에서 데이터를 가져오지 못했습니다.")
                return jsonify({
                    'success': True,
                    'message': 'ERPia 연결은 성공했지만 가져올 데이터가 없습니다.',
                    'data': {
                        'company_id': current_company_id,
                        'test_result': test_result,
                        'fetched_count': 0
                    }
                })
            
            logger.info(f"📊 총 {len(customers_data)}개 고객 데이터 수집 완료")
            
            # 데이터베이스 UPSERT 처리
            updated_count = 0
            inserted_count = 0
            error_count = 0
            
            def apply_classification_data(customer_data, classification_mapping, company_id):
                """ERPia 고객 데이터에 분류 정보 적용"""
                # 기본 분류 정보 초기화 (회사별로 다르게 처리 가능)
                classification_data = {
                    'distribution_type': None,  # 유통 (DIS)
                    'channel_type': None,       # 채널 (CH)
                    'sales_type': None,         # 매출 (SL)
                    'business_form': None,      # 매장형태 (TY)
                    'brand_zone': None,         # 브랜드존 (BZ)
                    'nuna_zoning': None,        # 뉴나 브랜드 조닝 (NZ)
                    'region': None,             # 지역 (RG)
                    'financial_group': None,    # 가결산 구분값 (FG)
                }
                
                # 회사별 기본 분류 설정
                if company_id == 1:  # 에이원
                    # 에이원 기본 분류 로직
                    if customer_data.get('business_type'):
                        if '도매' in customer_data['business_type']:
                            classification_data['distribution_type'] = 'WH'  # 도매
                        elif '소매' in customer_data['business_type']:
                            classification_data['distribution_type'] = 'RT'  # 소매
                    
                    # 거래처명 기반 채널 분류
                    customer_name = customer_data.get('customer_name', '').lower()
                    if '마트' in customer_name or '대형' in customer_name:
                        classification_data['channel_type'] = 'LM'  # 대형마트
                    elif '온라인' in customer_name or '쇼핑몰' in customer_name:
                        classification_data['channel_type'] = 'ON'  # 온라인
                    
                elif company_id == 2:  # 에이원월드
                    # 에이원월드 기본 분류 로직
                    classification_data['distribution_type'] = 'EX'  # 수출/해외
                
                # ERPia 데이터에 분류 정보 추가
                customer_data.update(classification_data)
                
                return customer_data
            
            # 디버깅을 위해 첫 번째 레코드만 처리
            test_customer = customers_data[0] if customers_data else None
            if test_customer:
                logger.info(f"🔍 첫 번째 고객 데이터 구조: {list(test_customer.keys())}")
                logger.info(f"🔍 첫 번째 고객 데이터 샘플: {dict(list(test_customer.items())[:5])}")
            
            for idx, customer_data in enumerate(customers_data):
                # 디버깅을 위해 처음 3개만 처리
                if idx >= 3:
                    break
                    
                try:
                    customer_code = customer_data.get('customer_code', '').strip()
                    if not customer_code:
                        logger.warning(f"⚠️ 고객 코드 없음: {customer_data}")
                        error_count += 1
                        continue
                    
                    logger.info(f"🔄 처리 중: {customer_code} - {customer_data.get('customer_name', 'N/A')}")
                    
                    # 분류 정보 적용
                    customer_data = apply_classification_data(customer_data, classification_mapping, current_company_id)
                    logger.info(f"📋 분류 적용 완료: {customer_code}")
                    
                    # 기존 고객 찾기
                    existing_customer = ErpiaCustomer.query.filter_by(
                        customer_code=customer_code,
                        company_id=current_company_id
                    ).first()
                    
                    if existing_customer:
                        # 업데이트
                        logger.info(f"📝 기존 고객 업데이트: {customer_code}")
                        
                        # 시스템 필드 제외하고 업데이트
                        system_fields = {
                            'seq', 'ins_user', 'ins_date', 'company_id'  # upt_user, upt_date는 수동 설정
                        }
                        
                        updated_fields = 0
                        for key, value in customer_data.items():
                            if hasattr(existing_customer, key) and key not in system_fields:
                                setattr(existing_customer, key, value)
                                updated_fields += 1
                            elif key in system_fields:
                                logger.debug(f"🔧 시스템 필드 건너뜀: {key}")
                            else:
                                logger.warning(f"⚠️ 알 수 없는 필드: {key}")
                        
                        logger.info(f"📋 업데이트 완료: {updated_fields}개 필드")
                        
                        existing_customer.upt_user = member_id
                        existing_customer.upt_date = datetime.utcnow()
                        updated_count += 1
                        
                    else:
                        # 신규 삽입
                        logger.info(f"➕ 신규 고객 추가: {customer_code}")
                        
                        # 시스템 필드 제외하고 필터링
                        system_fields = {
                            'seq', 'ins_user', 'ins_date', 'upt_user', 'upt_date', 'company_id'
                        }
                        
                        customer_data_filtered = {}
                        for key, value in customer_data.items():
                            if hasattr(ErpiaCustomer, key) and key not in system_fields:
                                customer_data_filtered[key] = value
                            elif key in system_fields:
                                logger.debug(f"🔧 시스템 필드 제외: {key}")
                            else:
                                logger.warning(f"⚠️ 모델에 없는 필드 제외: {key} = {value}")
                        
                        logger.info(f"📋 필터링 완료: {len(customer_data_filtered)}개 필드")
                        
                        new_customer = ErpiaCustomer(
                            company_id=current_company_id,
                            ins_user=member_id,
                            ins_date=datetime.utcnow(),
                            upt_user=member_id,
                            upt_date=datetime.utcnow(),
                            **customer_data_filtered
                        )
                        db.session.add(new_customer)
                        inserted_count += 1
                    
                    logger.info(f"✅ 고객 처리 완료: {customer_code}")
                    
                except Exception as e:
                    logger.error(f"❌ 고객 데이터 처리 실패: {customer_data.get('customer_code', 'Unknown')}")
                    logger.error(f"❌ 오류 상세: {str(e)}")
                    import traceback
                    logger.error(f"❌ 스택 트레이스: {traceback.format_exc()}")
                    error_count += 1
                    continue
            
            # 변경사항 커밋
            db.session.commit()
            
            result_message = f"동기화 완료: {updated_count}개 업데이트, {inserted_count}개 신규 추가"
            if error_count > 0:
                result_message += f", {error_count}개 오류"
            
            logger.info(f"✅ ERPia 매장 정보 동기화 완료: {result_message}")
            
            return jsonify({
                'success': True,
                'message': result_message,
                'data': {
                    'company_id': current_company_id,
                    'total_fetched': len(customers_data),
                    'updated': updated_count,
                    'inserted': inserted_count,
                    'errors': error_count,
                    'test_result': test_result
                }
            })
            
        except Exception as e:
            logger.error(f"❌ 데이터 동기화 실패: {str(e)}")
            return jsonify({'success': False, 'message': f'데이터 동기화 오류: {str(e)}'}), 500
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ ERPia 동기화 API 전체 오류: {str(e)}")
        import traceback
        logger.error(f"❌ 스택 트레이스: {traceback.format_exc()}")
        return jsonify({
            'success': False, 
            'message': f'서버 오류가 발생했습니다: {str(e)}'
        }), 500

@shop_bp.route('/api/export-excel')
def api_export_excel():
    """매장 정보 엑셀 다운로드 API - 동적 분류 완전 반영"""
    try:
        # 세션 체크
        if not session.get('member_seq'):
            return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401
        
        current_company_id = session.get('current_company_id', 1)
        shop_type = request.args.get('ty', 'All')
        search = request.args.get('search', '').strip()
        
        logger.info(f"📊 엑셀 다운로드 요청: company_id={current_company_id}, shop_type={shop_type}, search={search}")
        
        # 현재 동적 분류 정보 조회
        classifications = {}
        classification_headers = []
        
        # CST 그룹 찾기
        cst_group = Code.query.filter_by(code='CST', depth=0).first()
        if cst_group:
            # CST 하위의 모든 분류 그룹들을 동적으로 조회
            classification_groups = Code.query.filter_by(
                parent_seq=cst_group.seq, 
                depth=1
            ).order_by(Code.sort.asc()).all()
            
            for group in classification_groups:
                group_key = group.code.lower()
                classifications[group_key] = {
                    'code': group.code,
                    'name': group.code_name,
                    'field_name': f'{group_key}_type' if group_key not in ['dis', 'ch', 'sl', 'ty'] else {
                        'dis': 'distribution_type',
                        'ch': 'channel_type', 
                        'sl': 'sales_type',
                        'ty': 'business_form'
                    }.get(group_key, f'{group_key}_type')
                }
                classification_headers.append(group.code_name)
        
        # 기본 쿼리
        query = ErpiaCustomer.query.filter_by(company_id=current_company_id)
        
        # 매장 유형 필터링
        if shop_type == 'Shop':
            query = query.filter_by(shop_yn='Y')
        elif shop_type == 'NoShop':
            query = query.filter_by(shop_yn='N')
        elif shop_type == 'NotUsed':
            query = query.filter_by(shop_yn='NU')  # 매장사용안함
        elif shop_type == 'Nothing':
            query = query.filter(ErpiaCustomer.shop_yn.is_(None))
        
        # 검색 필터링
        if search:
            query = query.filter(
                or_(
                    ErpiaCustomer.customer_name.like(f'%{search}%'),
                    ErpiaCustomer.customer_code.like(f'%{search}%'),
                    ErpiaCustomer.our_manager.like(f'%{search}%'),
                    ErpiaCustomer.address1.like(f'%{search}%')
                )
            )
        
        # 데이터 조회
        shops = query.order_by(ErpiaCustomer.customer_name.asc()).all()
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "매장 정보"
        
        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 동적 헤더 구성 (기본 정보 + 동적 분류)
        base_headers = [
            "거래처코드", "거래처명", "대표자", "사업자번호", "업태", "종목",
            "전화번호", "팩스번호", "우리담당자", "상대방담당자", "상대방담당자전화", "상대방담당자전화2",
            "세금계산서우편번호", "세금계산서주소", "배송지우편번호", "배송지주소", 
            "로케이션"
        ]
        
        # 동적 분류 헤더 추가
        headers = base_headers + classification_headers + [
            "세금담당자", "세금담당자전화", "세금담당자이메일", "매장사용", 
            "비고", "메모", "등록일", "수정일"
        ]
        
        # 헤더 행 추가
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 데이터 행 추가
        for row_idx, shop in enumerate(shops, 2):
            # 기본 데이터
            base_data = [
                shop.customer_code or '',
                shop.customer_name or '',
                shop.ceo or '',
                shop.business_number or '',
                shop.business_type or '',
                shop.business_item or '',
                shop.phone or '',
                shop.fax or '',
                shop.our_manager or '',
                shop.customer_manager or '',
                shop.customer_manager_tel or '',
                shop.customer_manager_tel2 or '', # 새로운 필드 추가
                shop.zip_code1 or '',
                shop.address1 or '',
                shop.zip_code2 or '',
                shop.address2 or '',
                shop.location or ''
            ]
            
            # 동적 분류 데이터 추가
            classification_data = []
            for group_key, group_info in classifications.items():
                field_name = group_info['field_name']
                field_value = getattr(shop, field_name, '') or ''
                classification_data.append(field_value)
            
            # 나머지 데이터
            remaining_data = [
                shop.tax_manager or '',
                shop.tax_manager_tel or '',
                shop.tax_email or '',
                '매장' if shop.shop_yn == 'Y' else '매장아님' if shop.shop_yn == 'N' else '미분류',
                shop.remarks or '',
                shop.memo or '',
                shop.ins_date.strftime('%Y-%m-%d %H:%M') if shop.ins_date else '',
                shop.upt_date.strftime('%Y-%m-%d %H:%M') if shop.upt_date else ''
            ]
            
            # 전체 데이터 행 구성
            data_row = base_data + classification_data + remaining_data
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                
                # 매장사용 컬럼 색상 설정 (동적 위치 계산)
                shop_yn_col = len(base_headers) + len(classification_headers) + 4  # 매장사용 컬럼 위치
                if col == shop_yn_col:
                    if value == '매장':
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    elif value == '매장아님':
                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        # 두 번째 시트에 분류 코드 정보 추가 (업로드 참조용)
        ws_codes = wb.create_sheet("분류코드정보")
        ws_codes.cell(row=1, column=1, value="분류그룹").font = header_font
        ws_codes.cell(row=1, column=2, value="분류코드").font = header_font
        ws_codes.cell(row=1, column=3, value="분류명").font = header_font
        
        code_row = 2
        for group_key, group_info in classifications.items():
            # 각 분류 그룹의 하위 코드들 조회
            group = Code.query.filter_by(code=group_info['code'], depth=1).first()
            if group:
                sub_codes = Code.query.filter_by(parent_seq=group.seq, depth=2).order_by(Code.sort.asc()).all()
                for sub_code in sub_codes:
                    ws_codes.cell(row=code_row, column=1, value=group_info['name'])
                    ws_codes.cell(row=code_row, column=2, value=sub_code.code)
                    ws_codes.cell(row=code_row, column=3, value=sub_code.code_name)
                    code_row += 1
        
        # 컬럼 너비 자동 조정
        for sheet in [ws, ws_codes]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # 메모리 파일로 저장
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 파일명 생성
        from urllib.parse import quote
        company_name = "에이원" if current_company_id == 1 else "에이원월드"
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"매장정보_{company_name}_{timestamp}.xlsx"
        encoded_filename = quote(filename.encode('utf-8'))
        
        logger.info(f"✅ 엑셀 파일 생성 완료: {len(shops)}개 데이터, {len(classifications)}개 동적 분류, 파일명: {filename}")
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"❌ 엑셀 다운로드 실패: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'엑셀 생성 중 오류가 발생했습니다: {str(e)}'
        }), 500

@shop_bp.route('/api/upload-excel', methods=['POST'])
def api_upload_excel():
    """매장 정보 엑셀 일괄 업로드 API - 동적 분류 완전 지원"""
    try:
        # 세션 체크
        if not session.get('member_seq'):
            return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401
        
        current_company_id = session.get('current_company_id', 1)
        member_id = session.get('member_id', 'admin')
        
        logger.info(f"📤 엑셀 업로드 요청: company_id={current_company_id}, user={member_id}")
        
        # 파일 체크
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 선택되지 않았습니다.'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '파일이 선택되지 않았습니다.'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': '엑셀 파일만 업로드 가능합니다.'}), 400
        
        # 현재 동적 분류 정보 조회
        classifications = {}
        classification_field_map = {}  # 헤더명 -> 필드명 매핑
        
        # CST 그룹 찾기
        cst_group = Code.query.filter_by(code='CST', depth=0).first()
        if cst_group:
            classification_groups = Code.query.filter_by(
                parent_seq=cst_group.seq, 
                depth=1
            ).order_by(Code.sort.asc()).all()
            
            for group in classification_groups:
                group_key = group.code.lower()
                field_name = f'{group_key}_type' if group_key not in ['dis', 'ch', 'sl', 'ty'] else {
                    'dis': 'distribution_type',
                    'ch': 'channel_type', 
                    'sl': 'sales_type',
                    'ty': 'business_form'
                }.get(group_key, f'{group_key}_type')
                
                classifications[group_key] = {
                    'code': group.code,
                    'name': group.code_name,
                    'field_name': field_name
                }
                classification_field_map[group.code_name] = field_name
                
                # 분류 코드별 유효한 값들 조회
                sub_codes = Code.query.filter_by(parent_seq=group.seq, depth=2).all()
                valid_codes = {code.code: code.code_name for code in sub_codes}
                classifications[group_key]['valid_codes'] = valid_codes
        
        # 엑셀 파일 읽기
        try:
            wb = load_workbook(file.stream, read_only=True)
            ws = wb.active
        except Exception as e:
            return jsonify({'success': False, 'message': f'엑셀 파일을 읽을 수 없습니다: {str(e)}'}), 400
        
        # 헤더 행 읽기
        headers = []
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for cell in header_row:
            headers.append(str(cell) if cell is not None else '')
        
        # 필수 컬럼 체크
        required_columns = ['거래처코드', '거래처명']
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            return jsonify({
                'success': False, 
                'message': f'필수 컬럼이 누락되었습니다: {", ".join(missing_columns)}'
            }), 400
        
        # 데이터 처리
        success_count = 0
        update_count = 0
        insert_count = 0
        error_count = 0
        error_details = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # 행 데이터를 헤더와 매핑
                row_data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        header = headers[col_idx]
                        row_data[header] = str(value).strip() if value is not None else ''
                
                # 거래처코드 필수 체크
                customer_code = row_data.get('거래처코드', '').strip()
                if not customer_code:
                    error_details.append(f"행 {row_idx}: 거래처코드가 없습니다.")
                    error_count += 1
                    continue
                
                # 기존 데이터 찾기
                existing_customer = ErpiaCustomer.query.filter_by(
                    customer_code=customer_code,
                    company_id=current_company_id
                ).first()
                
                # 데이터 매핑
                update_data = {}
                
                # 기본 필드 매핑
                field_mapping = {
                    '거래처명': 'customer_name',
                    '대표자': 'ceo',
                    '사업자번호': 'business_number',
                    '업태': 'business_type',
                    '종목': 'business_item',
                    '전화번호': 'phone',
                    '팩스번호': 'fax',
                    '우리담당자': 'our_manager',
                    '상대방담당자': 'customer_manager',
                    '상대방담당자전화': 'customer_manager_tel',
                    '상대방담당자전화2': 'customer_manager_tel2', # 새로운 필드 추가
                    '세금계산서우편번호': 'zip_code1',
                    '세금계산서주소': 'address1',
                    '배송지우편번호': 'zip_code2',
                    '배송지주소': 'address2',
                    '로케이션': 'location',
                    '세금담당자': 'tax_manager',
                    '세금담당자전화': 'tax_manager_tel',
                    '세금담당자이메일': 'tax_email',
                    '비고': 'remarks',
                    '메모': 'memo'
                }
                
                for excel_col, db_field in field_mapping.items():
                    if excel_col in row_data and row_data[excel_col]:
                        update_data[db_field] = row_data[excel_col]
                
                # 매장사용 필드 처리
                if '매장사용' in row_data:
                    shop_status = row_data['매장사용'].strip()
                    if shop_status == '매장':
                        update_data['shop_yn'] = 'Y'
                    elif shop_status == '매장아님':
                        update_data['shop_yn'] = 'N'
                    else:
                        update_data['shop_yn'] = None
                
                # 동적 분류 필드 처리
                for header_name, field_name in classification_field_map.items():
                    if header_name in row_data and row_data[header_name]:
                        classification_value = row_data[header_name].strip()
                        
                        # 분류 코드 유효성 검증
                        group_key = next((k for k, v in classifications.items() 
                                        if v['name'] == header_name), None)
                        
                        if group_key and classification_value:
                            valid_codes = classifications[group_key]['valid_codes']
                            
                            # 코드명으로 입력된 경우 코드로 변환
                            if classification_value in valid_codes.values():
                                # 코드명에서 코드 찾기
                                for code, name in valid_codes.items():
                                    if name == classification_value:
                                        update_data[field_name] = code
                                        break
                            elif classification_value in valid_codes:
                                # 이미 코드로 입력된 경우
                                update_data[field_name] = classification_value
                            else:
                                # 유효하지 않은 분류값
                                error_details.append(f"행 {row_idx}: {header_name}의 값 '{classification_value}'은(는) 유효하지 않습니다.")
                                continue
                
                if existing_customer:
                    # 기존 데이터 업데이트
                    for field, value in update_data.items():
                        if hasattr(existing_customer, field):
                            setattr(existing_customer, field, value)
                    
                    existing_customer.upt_user = member_id
                    existing_customer.upt_date = datetime.utcnow()
                    update_count += 1
                    
                else:
                    # 신규 데이터 삽입
                    new_customer = ErpiaCustomer(
                        customer_code=customer_code,
                        company_id=current_company_id,
                        ins_user=member_id,
                        ins_date=datetime.utcnow(),
                        upt_user=member_id,
                        upt_date=datetime.utcnow(),
                        **update_data
                    )
                    db.session.add(new_customer)
                    insert_count += 1
                
                success_count += 1
                
            except Exception as e:
                error_details.append(f"행 {row_idx}: {str(e)}")
                error_count += 1
                continue
        
        # 변경사항 커밋
        try:
            db.session.commit()
            logger.info(f"✅ 엑셀 업로드 완료: 성공 {success_count}개 (업데이트 {update_count}, 삽입 {insert_count}), 오류 {error_count}개")
        except Exception as e:
            db.session.rollback()
            logger.error(f"❌ 엑셀 업로드 커밋 실패: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'데이터 저장 중 오류가 발생했습니다: {str(e)}'
            }), 500
        
        result_message = f"업로드 완료: {success_count}개 성공 (업데이트 {update_count}개, 신규 {insert_count}개)"
        if error_count > 0:
            result_message += f", {error_count}개 오류"
        
        return jsonify({
            'success': True,
            'message': result_message,
            'data': {
                'total_processed': success_count + error_count,
                'success_count': success_count,
                'update_count': update_count,
                'insert_count': insert_count,
                'error_count': error_count,
                'error_details': error_details[:10]  # 최대 10개만 반환
            }
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ 엑셀 업로드 실패: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'업로드 중 오류가 발생했습니다: {str(e)}'
        }), 500 